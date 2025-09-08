
# chart_widget.py
import io
import platform
import tempfile
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import Optional
 
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.ticker as mticker
from matplotlib.ticker import ScalarFormatter, MaxNLocator
 
DEFAULT_OPTIONS = {
    "title": "P&L Chart",
    "xlabel": "X",
    "ylabel": "Y",
    "show_grid": True,
    "show_legend": True,
    "y_commas": True,
    "x_ticks": 5,         # number of x ticks
    "y_ticks": None,      # None = let mpl pick
    "margins": (0.05, 0.20),
    # physical sizing (in pixels) and DPI
    "width_px": None,
    "height_px": None,
    "dpi": 100,
    "custom_size": False,
    "min_width_px": 600,
    "min_height_px": 400,
    # x-axis integer tick option
    "x_integer_ticks": False,
    # optional vertical reference line
    "ref_line": False,
    "ref_x": None,
    "ref_color": "#51284F",
    "ref_style": "-.",
    "ref_width": 1.25,
    "ref_alpha": 0.9,
    # symmetric x-limits options
    "symmetric_x": False,        # keep x-limits symmetric around a center
    "center_mode": "auto",     # one of: "auto"|"ref"|"x_ref"|"value"|"zero"
    "center_value": None,        # used when center_mode == "value"
    # style / theme hooks
    "label_style": "Title.TLabel",
    "dropdown_style": "TCombobox",
    "text_style": "TLabel",
    "main_color": "#51284F",
    "accent_color": "#0078D4",
    # stats panel options
    "show_line_stats": True,   # show per-line stats panel
    "show_max_in_summary": True,     # draw a marker at each line's max
    # customizable UI labels (for Customize dialog)
    "label_ref_line": "Reference line",
    "label_show_line_stats": "Show line statistics",
    "label_show_max_in_summary": "Display max in summary",
    "max_statistic_label" : "Max_Label",
    "label_x_cross" : "X-Cross",
    "show_custom_message": False,
    "label_show_custom_message": "Show custom message",
    # extra padding
    "extra_bottom_pad": 40,   # additional white space under the axes (0..~0.4)
    "custom_message": "",  # optional multi-line message rendered under stats
}
 
class ChartWidget(ttk.Frame):
    """
    Minimal, reusable chart widget:
      - set_data(x, series)
      - set_options(**opts) / update_options(dict)
      - refresh()
 
    x: list[float]
    series: dict[str, list[float]]  # curves keyed by label
    options: see DEFAULT_OPTIONS
    """
    def __init__(self, parent, *, options=None):
        super().__init__(parent)
        self.options = dict(DEFAULT_OPTIONS)
        if options:
            self.options.update(options)
 
        self._x = []               # list[float]
        self._series = {}          # dict[str, list[float]]
 
        # Optional external refresh callback (lets parent recompute before drawing)
        self._external_refresh = None
        try:
            cb = (self.options.get("refresh_callback") if isinstance(self.options, dict) else None)
            if cb and callable(cb):
                self._external_refresh = cb
        except Exception:
            pass

        # header row (optional small toolbar)
        top = ttk.Frame(self)
        top.pack(fill="x", pady=(0, 6))
        self._title_var = tk.StringVar(value=self.options.get("title", "Chart"))
        self._title_lbl = ttk.Label(top, textvariable=self._title_var, style=self.options.get("label_style", "Title.TLabel"))
        self._title_lbl.pack(side="left")
        ttk.Button(top, text="Customize…", command=self._open_customize_window).pack(side="left", padx=(6, 0))
        ttk.Button(top, text="Refresh", command=self._on_refresh).pack(side="right", padx=(6, 0))
        ttk.Button(top, text="Copy Table", command=self.copy_table_to_clipboard).pack(side="right", padx=(6, 0))
        ttk.Button(top, text="Copy Chart", command=self.copy_chart_to_clipboard).pack(side="right", padx=(6, 0))
        ttk.Button(top, text="Export Excel", command=self.export_to_excel).pack(side="right")
 
        # mpl canvas
        dpi = int(self.options.get("dpi", 100) or 100)
        self._fig = Figure(figsize=(5, 3), dpi=dpi)
        self._ax = self._fig.add_subplot(111)
        self._canvas = FigureCanvasTkAgg(self._fig, master=self)
        self._canvas_widget = self._canvas.get_tk_widget()
        # apply initial sizing based on width_px/height_px options
        self._apply_size_from_options()
        self._canvas.draw()
        # default pack policy (can be overridden by custom_size)
        try:
            self._canvas_widget.pack_forget()
        except Exception:
            pass
        self._canvas_widget.pack(fill="both", expand=True)
 
        # on-axes stats text handle (drawn below x-axis)
        self._stats_text = None
 
        # keep custom-sized canvas fixed on parent resizes
        try:
            self.bind("<Configure>", self._on_parent_configure)
        except Exception:
            pass
 
        self._draw_placeholder("No data yet — call set_data(...)")

    def _on_refresh(self):
        """If an external refresh callback is provided, call it; otherwise refresh current data."""
        try:
            if callable(self._external_refresh):
                self._external_refresh()
                return
        except Exception:
            pass
        self.refresh()
 
    # ---------- Public API ----------
    def set_data(self, x, series: dict):
        self._x = list(x or [])
        self._series = dict(series or {})
        return self
 
    def set_options(self, **opts):
        self.options.update(opts)
        self._sync_header_from_options()
        self._apply_size_from_options()
        try:
            self.refresh()
        except Exception:
            pass
        return self
 
    def update_options(self, opts: dict):
        if opts:
            self.options.update(opts)
            self._sync_header_from_options()
            self._apply_size_from_options()
            try:
                self.refresh()
            except Exception:
                pass
        return self
 
    def _sync_header_from_options(self):
        """Sync toolbar title from self.options (public API-updatable)."""
        try:
            if hasattr(self, "_title_var"):
                self._title_var.set(self.options.get("title", "Chart"))
        except Exception:
            pass
 
    def set_x_center(self, mode: str = "auto", value: Optional[float] = None, symmetric: bool = True):
        """Public API to control x-centering.
 
        mode: one of {"auto", "zero", "value", "ref", "x_ref"}  (note: "x_ref" behaves the same as "ref")
          - "auto": natural middle of data
          - "zero": center at x=0
          - "value": center at provided `value`
          - "ref": center at `options['ref_x']` if set
        value: used only when mode == "value"
        symmetric: whether to enforce symmetric x-limits around the center
        """
        try:
            mode = str(mode).lower()
            if mode not in {"auto", "zero", "value", "ref", "x_ref"}:
                mode = "auto"
            if mode == "x_ref":
                mode = "ref"
            self.options["center_mode"] = mode
            self.options["center_value"] = float(value) if (mode == "value" and value is not None) else None
            self.options["symmetric_x"] = bool(symmetric)
            # Apply immediately
            self.refresh()
        except Exception:
            pass
 
    def _apply_pack_policy(self):
        """Apply pack/grid behavior depending on custom_size option."""
        try:
            cs = bool(self.options.get("custom_size", False))
            if cs:
                # prevent the frame from resizing to its children
                try:
                    self.pack_propagate(False)
                except Exception:
                    pass
                try:
                    self.grid_propagate(False)
                except Exception:
                    pass
                # fix the canvas' packing so it doesn't expand
                try:
                    self._canvas_widget.pack_forget()
                except Exception:
                    pass
                self._canvas_widget.pack(fill=None, expand=False)
            else:
                # allow responsive behavior
                try:
                    self.pack_propagate(True)
                except Exception:
                    pass
                try:
                    self.grid_propagate(True)
                except Exception:
                    pass
                try:
                    self._canvas_widget.pack_forget()
                except Exception:
                    pass
                self._canvas_widget.pack(fill="both", expand=True)
        except Exception:
            pass
 
    def _apply_size_from_options(self):
        """Apply physical size (in px) and dpi from options to the Figure and Tk canvas."""
        try:
            dpi = int(self.options.get("dpi", 100) or 100)
            self._fig.set_dpi(dpi)
        except Exception:
            dpi = 100
        # default: no fixed size
        self._fixed_size = None
        try:
            wpx = self.options.get("width_px", None)
            hpx = self.options.get("height_px", None)
            min_wpx = int(self.options.get("min_width_px", 600) or 600)
            min_hpx = int(self.options.get("min_height_px", 400) or 400)
            if wpx is not None and hpx is not None:
                wpx = int(wpx)
                hpx = int(hpx)
                # Clamp to minimums
                wpx = max(min_wpx, wpx)
                hpx = max(min_hpx, hpx)
                # px -> inches
                win = max(1, wpx) / float(dpi)
                hin = max(1, hpx) / float(dpi)
                self._fig.set_size_inches(win, hin, forward=True)
                # fix the Tk canvas widget size to requested pixels
                try:
                    self._canvas_widget.configure(width=wpx, height=hpx)
                except Exception:
                    try:
                        self._canvas.get_tk_widget().configure(width=wpx, height=hpx)
                    except Exception:
                        pass
                self._fixed_size = (wpx, hpx)
        except Exception:
            pass
        # apply pack policy depending on custom_size
        self._apply_pack_policy()
 
    def _on_parent_configure(self, *_):
        """Keep the canvas size fixed when custom_size is enabled."""
        try:
            if not bool(self.options.get("custom_size", False)):
                return
            if not getattr(self, "_fixed_size", None):
                return
            wpx, hpx = self._fixed_size
            # Clamp to minimums
            min_wpx = int(self.options.get("min_width_px", 600) or 600)
            min_hpx = int(self.options.get("min_height_px", 400) or 400)
            wpx = max(min_wpx, wpx)
            hpx = max(min_hpx, hpx)
            # clamp the canvas to fixed pixel size
            try:
                self._canvas_widget.configure(width=wpx, height=hpx)
            except Exception:
                pass
            try:
                dpi = float(self._fig.get_dpi() or 100.0)
                self._fig.set_size_inches(wpx / dpi, hpx / dpi, forward=True)
            except Exception:
                pass
        except Exception:
            pass
 
    def _zero_crossings(self, x, y):
        """Return list of x-values where the series crosses y=0 (linear interpolation)."""
        xs = []
        n = min(len(x), len(y))
        if n == 0:
            return xs
        # exact zeros
        for i in range(n):
            try:
                if float(y[i]) == 0.0:
                    xs.append(float(x[i]))
            except Exception:
                pass
        # sign changes
        for i in range(n - 1):
            try:
                y1 = float(y[i]); y2 = float(y[i+1])
                if y1 == 0.0 or y2 == 0.0:
                    continue
                if (y1 > 0 and y2 < 0) or (y1 < 0 and y2 > 0):
                    x1 = float(x[i]); x2 = float(x[i+1])
                    # linear interpolation for y=0
                    t = -y1 / (y2 - y1)
                    x0 = x1 + t * (x2 - x1)
                    xs.append(x0)
            except Exception:
                continue
        # deduplicate-ish while keeping order
        out = []
        seen = set()
        for v in xs:
            key = round(v, 9)
            if key not in seen:
                seen.add(key)
                out.append(v)
        return out
 
    def refresh(self):
        if not self._x or not self._series:
            self._draw_placeholder("No curves to display")
            return
       
        # Re-assert fixed canvas size when custom_size is active
        try:
            if bool(self.options.get("custom_size", False)) and getattr(self, "_fixed_size", None):
                wpx, hpx = self._fixed_size
                self._canvas_widget.configure(width=wpx, height=hpx)
                dpi = float(self._fig.get_dpi() or 100.0)
                self._fig.set_size_inches(wpx / dpi, hpx / dpi, forward=True)
        except Exception:
            pass
 
        self._ax.clear()
 
        # plot all series and keep handles/colors for markers
        _plotted = []  # list of (label, line2d, yvals)
        for label, yvals in self._series.items():
            if not yvals:
                continue
            line, = self._ax.plot(self._x, yvals, label=label)
            _plotted.append((label, line, yvals))
 
        # zero line (handy for P&L)
        self._ax.axhline(0.0, linewidth=1, color="black")
 
        # titles/labels
        self._ax.set_title(self.options.get("title", ""))
        self._ax.set_xlabel(self.options.get("xlabel", ""))
        self._ax.set_ylabel(self.options.get("ylabel", ""))
 
        # reference line
        if self.options.get("ref_line") and self.options.get("ref_x") is not None:
            self._ax.axvline(
                x=float(self.options["ref_x"]),
                color=self.options.get("ref_color") or self.options.get("main_color", "#51284F"),
                linewidth=float(self.options.get("ref_width", 1.25)),
                linestyle=str(self.options.get("ref_style", "-.")),
                alpha=float(self.options.get("ref_alpha", 0.9)),
                zorder=10,
                label=self.options.get("ref_label") if self.options.get("show_legend", True) else None
            )
 
        # legend / grid
        if self.options.get("show_legend", True):
            self._ax.legend(loc="best")
        if self.options.get("show_grid", True):
            self._ax.grid(True, linestyle="--", alpha=0.6)
        else:
            self._ax.grid(False)
 
        # tick formatting
        if self.options.get("y_commas", True):
            self._ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, pos: f"{v:,.0f}"))
        else:
            self._ax.yaxis.set_major_formatter(ScalarFormatter(useOffset=False))
            self._ax.ticklabel_format(style='plain', axis='y')
 
        # x granularity and integer ticks
        xt = int(self.options.get("x_ticks") or 0)
        x_integer = bool(self.options.get("x_integer_ticks", False))
        if xt and len(self._x) >= 2:
            xmin, xmax = min(self._x), max(self._x)
            if xmax > xmin:
                step = (xmax - xmin) / (xt - 1)
                ticks = [xmin + i * step for i in range(xt)]
                if x_integer:
                    ticks = sorted({int(round(t)) for t in ticks})
                self._ax.set_xticks(ticks)
        elif x_integer:
            # No explicit xt count; ask mpl to use integer locators
            self._ax.xaxis.set_major_locator(MaxNLocator(integer=True))
 
        # y granularity (optional)
        yt = self.options.get("y_ticks")
        if isinstance(yt, int) and yt >= 2:
            ys = [v for vals in self._series.values() for v in vals if isinstance(vals, (list, tuple))]
            flat = []
            for vals in self._series.values():
                flat.extend(vals or [])
            if flat:
                ymin, ymax = min(flat), max(flat)
                if ymax > ymin:
                    step = (ymax - ymin) / (yt - 1)
                    self._ax.set_yticks([ymin + i * step for i in range(yt)])
 
        # optional symmetric x-limits around a center value
        if self.options.get("symmetric_x", False) and len(self._x) >= 2:
            xmin, xmax = min(self._x), max(self._x)
            # determine center
            mode = str(self.options.get("center_mode", "auto")).lower()
            center = None
            if mode in ("ref", "x_ref"):
                cx = self.options.get("ref_x", None)
                center = float(cx) if cx is not None else None
            elif mode == "value":
                cv = self.options.get("center_value", None)
                center = float(cv) if cv is not None else None
            elif mode == "zero":
                center = 0.0
            # fallback to natural middle if no usable center
            if center is None:
                center = 0.5 * (xmin + xmax)
            # choose half-range to include all points
            half = max(abs(center - xmin), abs(xmax - center))
            # avoid degenerate range
            if half <= 0:
                half = max(1.0, 0.5 * (abs(xmin) + abs(xmax) + 1.0))
            self._ax.set_xlim(center - half, center + half)
 
        # margins + layout
        mx, my = self.options.get("margins", (0.05, 0.20))
        self._ax.margins(x=float(mx), y=float(my))
        self._fig.tight_layout(pad=2.0)
 
        # ---- per-line statistics panel (drawn in figure coords, beneath axes) ----
        # Clear any prior stats text
        try:
            if getattr(self, "_stats_text", None) is not None:
                self._stats_text.remove()
                self._stats_text = None
        except Exception:
            pass
        # Clear any prior custom message text
        try:
            if getattr(self, "_custom_text", None) is not None:
                self._custom_text.remove()
                self._custom_text = None
        except Exception:
            pass
 
        # Determine if stats are to be shown and compute lines
        show_stats = bool(self.options.get("show_line_stats", False))
        lines = []
        if show_stats:
            for label, line, yvals in _plotted:
                try:
                    xz = self._zero_crossings(self._x, yvals)
                    xz_txt = ", ".join(f"{v:,.2f}" for v in xz) if xz else "—"

                    if self.options.get("show_max_in_summary", True):
                        # Max
                        y_max = max(yvals)
                        idx_max = yvals.index(y_max)
                        x_at_max = self._x[idx_max] if idx_max < len(self._x) else None
                        label_max = str(self.options.get("max_statistic_label", "Max"))
                        # Min
                        y_min = min(yvals)
                        idx_min = yvals.index(y_min)
                        x_at_min = self._x[idx_min] if idx_min < len(self._x) else None
                        label_min = "Min"

                        label_cross = str(self.options.get("label_x_cross", "X-Cross"))
                        # Build only max/min parts; do NOT include the X-Cross here to avoid duplication
                        parts = []
                        if x_at_max is not None:
                            parts.append(f"{label_max} = {y_max:,.0f} at x={x_at_max:,.2f}")
                        else:
                            parts.append(f"{label_max} = {y_max:,.0f}")
                        if x_at_min is not None:
                            parts.append(f"{label_min} = {y_min:,.0f} at x={x_at_min:,.2f}")
                        else:
                            parts.append(f"{label_min} = {y_min:,.0f}")
                        suffix = ("  |  " + "  |  ".join(parts)) if parts else ""
                    else:
                        label_cross = str(self.options.get("label_x_cross", "X-Cross"))
                        suffix = ""
                    lines.append(f"{label}:  {label_cross} = {xz_txt}{suffix}")
                except Exception:
                    continue
 
        # Optional custom message (multi-line) to render under stats
        show_custom_msg = bool(self.options.get("show_custom_message", False))
        custom_msg = str(self.options.get("custom_message", "") or "")
        custom_lines = [ln for ln in custom_msg.splitlines() if ln.strip()] if show_custom_msg else []
 
        # Compute an ample bottom margin for stats and custom message to avoid any overlap
        n_lines = len(lines) if show_stats else 0
        m_lines = len(custom_lines) if show_custom_msg else 0
        bottom_needed = 0.1
        if n_lines:
            bottom_needed += 0.05 * max(0, n_lines)
        if m_lines:
            bottom_needed += 0.05 * max(0, m_lines)
        # bottom_needed = min(0.25, bottom_needed)
        print(f"bottom_needed:{bottom_needed}")
       
        # Apply bottom margin after tight_layout, so there's always space for the stats and custom message
        try:
            if bottom_needed == 0.1:
                self._fig.subplots_adjust(bottom=bottom_needed+0.05)
            else:
                if m_lines:
                    self._fig.subplots_adjust(bottom=bottom_needed-0.1)
                else:
                    print(f"in else")
                    self._fig.subplots_adjust(bottom=bottom_needed+0.03)
        except Exception:
            print(f"Failed to Set Bottom")
            pass
 
        # Place stats and optional custom message within the reserved bottom margin (figure coords)
        try:
            ax_pos = self._ax.get_position()
            print(ax_pos.y0)
            left = ax_pos.x0
            print(f"bottom axis: {ax_pos.y0}")
        except Exception:
            left = 0.1
 
        # Choose vertical anchors inside the reserved margin
        # If both present, put stats higher and custom message lower.
        if show_stats:
            stats_str = "\n".join(lines) if lines else ""
            try:
                stats_y = bottom_needed * (0.65 if (custom_lines) else 0.50)
                print(f"stats_y: {stats_y}")
                self._stats_text = self._fig.text(
                    left, stats_y, stats_str,
                    ha="left", va="top",
                    fontsize=9, color="gray", wrap=True,
                )
            except Exception:
                self._stats_text = None
        # Draw custom message under stats (or alone if no stats)
        if custom_lines:
            msg_str = "\n".join(custom_lines)
            try:
                msg_y = bottom_needed * ((0.50 if (show_stats and lines) else 0.6)) #- (0.05 * m_lines)
                print(f"msg_y: {msg_y}")
                self._custom_text = self._fig.text(
                    left, msg_y, msg_str,
                    ha="left", va="top",
                    fontsize=9, color="gray", wrap=True,
                )
            except Exception:
                self._custom_text = None
 
        self._canvas.draw()
 
    # ---------- Utilities ----------
    def copy_chart_to_clipboard(self):
        """Copy the current chart image to clipboard (Windows) or save a temp PNG (macOS/Linux)."""
        try:
            buf = io.BytesIO()
            self._fig.savefig(buf, format="png", dpi=200, bbox_inches="tight")
            data_png = buf.getvalue()
        except Exception as e:
            messagebox.showerror("Copy Chart", f"Failed to render chart:\n{e}")
            return
 
        if platform.system() == "Windows":
            try:
                from PIL import Image
                import win32clipboard, win32con
                buf.seek(0)
                img = Image.open(buf).convert("RGB")
                out = io.BytesIO()
                img.save(out, format="BMP")
                dib = out.getvalue()[14:]  # strip BMP header
                win32clipboard.OpenClipboard()
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardData(win32con.CF_DIB, dib)
                win32clipboard.CloseClipboard()
                messagebox.showinfo("Copy Chart", "Chart image copied to clipboard.")
                return
            except Exception:
                pass
 
        # fallback: write temp PNG
        try:
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            tmp.write(data_png); tmp.flush(); tmp.close()
            messagebox.showinfo("Copy Chart", f"Saved chart image to:\n{tmp.name}\n(Clipboard copy not available on this platform.)")
        except Exception as e:
            messagebox.showerror("Copy Chart", f"Failed to save chart image:\n{e}")
 
    def copy_table_to_clipboard(self):
        """Copy the current (x, series...) grid as TSV."""
        if not self._x or not self._series:
            messagebox.showwarning("Copy Table", "Nothing to copy.")
            return
        from io import StringIO
        sio = StringIO()
        labels = list(self._series.keys())
        sio.write("\t".join(["X"] + labels) + "\n")
        n = len(self._x)
        for i in range(n):
            row = [f"{self._x[i]:.6f}"]
            for lab in labels:
                vals = self._series.get(lab, [])
                row.append("" if i >= len(vals) else f"{vals[i]:.6f}")
            sio.write("\t".join(row) + "\n")
        tsv = sio.getvalue()
        try:
            self.clipboard_clear()
            self.clipboard_append(tsv)
            self.update()
            messagebox.showinfo("Copy Table", "Table copied to clipboard (TSV).")
        except Exception as e:
            messagebox.showerror("Copy Table", f"Clipboard copy failed:\n{e}")
 
    def export_to_excel(self, path: Optional[str] = None):
        if not self._x or not self._series:
            messagebox.showwarning("Export", "Nothing to export.")
            return
        if path is None:
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Workbook", "*.xlsx")],
                title="Save Table as Excel"
            )
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import numbers
        except Exception:
            messagebox.showerror("Export", "openpyxl is not installed.\n\npip install openpyxl")
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "ChartData"
            labels = list(self._series.keys())
            ws.append(["X"] + labels)
            n = len(self._x)
            for i in range(n):
                row = [self._x[i]]
                for lab in labels:
                    vals = self._series.get(lab, [])
                    row.append(vals[i] if i < len(vals) else None)
                ws.append(row)
            # basic formatting
            for cell in ws["A"][1:]:
                cell.number_format = numbers.FORMAT_NUMBER_00
            for col in range(2, 2 + len(labels)):
                for cell in ws[get_column_letter(col)][1:]:
                    cell.number_format = numbers.FORMAT_NUMBER_00
            # autosize
            for col in range(1, ws.max_column + 1):
                letter = get_column_letter(col)
                ws.column_dimensions[letter].width = max(12, min(40, max(len(str(c.value)) for c in ws[letter]) + 2))
            wb.save(path)
            messagebox.showinfo("Export", f"Saved to:\n{path}")
        except Exception as e:
            messagebox.showerror("Export", f"Failed to write Excel:\n{e}")
 
    def _open_customize_window(self):
        """Pop out a persistent customization window for chart options."""
        # Prevent multiple windows
        if getattr(self, "_customize_win", None) and tk.Toplevel.winfo_exists(self._customize_win):
            try:
                self._customize_win.deiconify()
                self._customize_win.lift()
                self._customize_win.focus_force()
            except Exception:
                pass
            return
 
        win = tk.Toplevel(self)
        self._customize_win = win
        win.title("Customize Chart")
 
        frm = ttk.Frame(win, padding=10)
        frm.pack(fill="both", expand=True)
 
        # Helpers to get/set vars from current options
        def opt_get(key, default=None):
            return self.options.get(key, default)
 
        # ----- Section: General -----
        ttk.Label(frm, text="General", style=self.options.get("text_style", "TLabel")).grid(row=0, column=0, sticky="w", pady=(0,4))
 
        ttk.Label(frm, text="Title:").grid(row=1, column=0, sticky="w")
        title_var = tk.StringVar(value=str(opt_get("title", "P&L Chart")))
        ttk.Entry(frm, textvariable=title_var, width=28).grid(row=1, column=1, sticky="w", padx=(6,0))
 
        ttk.Label(frm, text="X Label:").grid(row=2, column=0, sticky="w")
        xlabel_var = tk.StringVar(value=str(opt_get("xlabel", "X")))
        ttk.Entry(frm, textvariable=xlabel_var, width=28).grid(row=2, column=1, sticky="w", padx=(6,0))
 
        ttk.Label(frm, text="Y Label:").grid(row=3, column=0, sticky="w")
        ylabel_var = tk.StringVar(value=str(opt_get("ylabel", "Y")))
        ttk.Entry(frm, textvariable=ylabel_var, width=28).grid(row=3, column=1, sticky="w", padx=(6,0))
 
        ttk.Label(frm, text="X Granularity:").grid(row=4, column=0, sticky="w")
        x_ticks_var = tk.StringVar(value=str(opt_get("x_ticks", 5)))
        x_ticks_ent = ttk.Entry(frm, textvariable=x_ticks_var, width=8)
        x_ticks_ent.grid(row=4, column=1, sticky="w", padx=(6,0))
 
        ttk.Label(frm, text="Y Granularity:").grid(row=5, column=0, sticky="w")
        y_ticks_val = opt_get("y_ticks", None)
        y_ticks_var = tk.StringVar(value="" if y_ticks_val in (None, "None") else str(y_ticks_val))
        y_ticks_ent = ttk.Entry(frm, textvariable=y_ticks_var, width=8)
        y_ticks_ent.grid(row=5, column=1, sticky="w", padx=(6,0))
 
        ttk.Label(frm, text="Chart Margins (x, y):").grid(row=6, column=0, sticky="w")
        mx, my = opt_get("margins", (0.05, 0.20)) or (0.05, 0.20)
        margin_x_var = tk.StringVar(value=str(mx))
        margin_y_var = tk.StringVar(value=str(my))
        mrow = ttk.Frame(frm)
        mrow.grid(row=6, column=1, sticky="w")
        ttk.Entry(mrow, textvariable=margin_x_var, width=8).pack(side="left")
        ttk.Label(mrow, text=",").pack(side="left", padx=2)
        ttk.Entry(mrow, textvariable=margin_y_var, width=8).pack(side="left")
 
        ttk.Separator(frm).grid(row=7, column=0, columnspan=2, sticky="ew", pady=(8,8))
 
        # ----- Section: Size -----
        ttk.Label(frm, text="Size", style=self.options.get("text_style", "TLabel")).grid(row=8, column=0, sticky="w", pady=(0,4))
 
        custom_size_var = tk.BooleanVar(value=bool(self.options.get("custom_size", False)))
        ttk.Checkbutton(frm, text="Use Custom Fixed Size", variable=custom_size_var).grid(row=9, column=0, sticky="w")
 
        ttk.Label(frm, text="Width (px):").grid(row=10, column=0, sticky="w")
        width_px_var = tk.StringVar(value="" if self.options.get("width_px") in (None, "None") else str(self.options.get("width_px")))
        width_px_ent = ttk.Entry(frm, textvariable=width_px_var, width=10)
        width_px_ent.grid(row=10, column=1, sticky="w", padx=(6,0))
 
        ttk.Label(frm, text="Height (px):").grid(row=11, column=0, sticky="w")
        height_px_var = tk.StringVar(value="" if self.options.get("height_px") in (None, "None") else str(self.options.get("height_px")))
        height_px_ent = ttk.Entry(frm, textvariable=height_px_var, width=10)
        height_px_ent.grid(row=11, column=1, sticky="w", padx=(6,0))
 
        ttk.Label(frm, text="DPI:").grid(row=12, column=0, sticky="w")
        dpi_var = tk.StringVar(value=str(self.options.get("dpi", 100)))
        ttk.Entry(frm, textvariable=dpi_var, width=10).grid(row=12, column=1, sticky="w", padx=(6,0))
 
        # --- Helper: Toggle width/height editability based on custom_size_var ---
        def _toggle_size_entries(*_):
            try:
                if bool(custom_size_var.get()):
                    # enable editing
                    try:
                        width_px_ent.state(["!disabled", "!readonly"])
                        height_px_ent.state(["!disabled", "!readonly"])
                    except Exception:
                        pass
                else:
                    # clear and make readonly
                    try:
                        width_px_var.set("")
                        height_px_var.set("")
                    except Exception:
                        pass
                    try:
                        width_px_ent.state(["readonly"])
                        height_px_ent.state(["readonly"])
                    except Exception:
                        pass
            except Exception:
                pass
 
        # tie to checkbox and set initial state
        custom_size_var.trace_add("write", _toggle_size_entries)
        _toggle_size_entries()
 
        ttk.Separator(frm).grid(row=13, column=0, columnspan=2, sticky="ew", pady=(8,8))
 
        # ----- Section: Format -----
        ttk.Label(frm, text="Format", style=self.options.get("text_style", "TLabel")).grid(row=14, column=0, sticky="w", pady=(0,4))
 
        show_grid_var = tk.BooleanVar(value=bool(opt_get("show_grid", True)))
        show_legend_var = tk.BooleanVar(value=bool(opt_get("show_legend", True)))
        y_commas_var = tk.BooleanVar(value=bool(opt_get("y_commas", True)))
        ttk.Checkbutton(frm, text="Show Grid", variable=show_grid_var).grid(row=15, column=0, sticky="w")
        ttk.Checkbutton(frm, text="Show Legend", variable=show_legend_var).grid(row=15, column=1, sticky="w")
        ttk.Checkbutton(frm, text="Commas", variable=y_commas_var).grid(row=16, column=0, sticky="w")
 
        xint_var = tk.BooleanVar(value=bool(opt_get("x_integer_ticks", False)))
        ttk.Checkbutton(frm, text="Force integer X ticks", variable=xint_var).grid(row=16, column=1, sticky="w")
 
        sym_var = tk.BooleanVar(value=bool(opt_get("symmetric_x", False)))
        ttk.Checkbutton(frm, text="Symmetric X Range", variable=sym_var).grid(row=17, column=0, sticky="w", pady=(6,0))
 
        ttk.Label(frm, text="Center Around:").grid(row=18, column=0, sticky="w")
        center_mode_var = tk.StringVar(value=str(opt_get("center_mode", "auto")))
        mode_row = ttk.Frame(frm)
        mode_row.grid(row=18, column=1, sticky="w")
        ttk.Radiobutton(mode_row, text="Natural middle", value="auto", variable=center_mode_var).pack(side="left")
        ttk.Radiobutton(mode_row, text="Zero", value="zero", variable=center_mode_var).pack(side="left", padx=(8,0))
        ttk.Radiobutton(mode_row, text=opt_get("label_ref_line","Ref"), value="x_ref", variable=center_mode_var).pack(side="left", padx=(8,0))
        ttk.Radiobutton(mode_row, text="Custom", value="value", variable=center_mode_var).pack(side="left", padx=(8,0))
 
        ttk.Label(frm, text="Custom Center:").grid(row=19, column=0, sticky="w")
        center_val = opt_get("center_value", None)
        center_val_var = tk.StringVar(value="" if center_val in (None, "None") else str(center_val))
        center_ent = ttk.Entry(frm, textvariable=center_val_var, width=16)
        center_ent.grid(row=19, column=1, sticky="w", padx=(6,0))
 
        # enable/disable entry based on mode
        def _toggle_center_entry(*_):
            if center_mode_var.get() == "value":
                center_ent.state(["!disabled"])
            else:
                center_ent.state(["disabled"])
        _toggle_center_entry()
        center_mode_var.trace_add("write", _toggle_center_entry)
 
        ttk.Label(frm, text="Main Color (hex):").grid(row=20, column=0, sticky="w")
        main_color_var = tk.StringVar(value=str(opt_get("main_color", "#51284F")))
        ttk.Entry(frm, textvariable=main_color_var, width=16).grid(row=20, column=1, sticky="w", padx=(6,0))
 
        ttk.Label(frm, text="Accent Color (hex):").grid(row=21, column=0, sticky="w")
        accent_color_var = tk.StringVar(value=str(opt_get("accent_color", "#0078D4")))
        ttk.Entry(frm, textvariable=accent_color_var, width=16).grid(row=21, column=1, sticky="w", padx=(6,0))
 
        ttk.Separator(frm).grid(row=22, column=0, columnspan=2, sticky="ew", pady=(8,8))
 
        # ----- Section: Reference Line -----
        ttk.Label(frm, text=self.options.get("label_ref_line", "Reference line"), style=self.options.get("text_style", "TLabel")).grid(row=23, column=0, sticky="w", pady=(0,4))
 
        ref_line_var = tk.BooleanVar(value=bool(opt_get("ref_line", False)))
        ttk.Checkbutton(frm, text=self.options.get("label_ref_line", "Reference line"), variable=ref_line_var).grid(row=24, column=0, sticky="w")
 
        ttk.Label(frm, text="Ref X:").grid(row=25, column=0, sticky="w")
        ref_x_var = tk.StringVar(value="" if opt_get("ref_x", None) in (None, "None") else str(opt_get("ref_x")))
        ttk.Entry(frm, textvariable=ref_x_var, width=12).grid(row=25, column=1, sticky="w", padx=(6,0))
 
        ttk.Label(frm, text="Style:").grid(row=26, column=0, sticky="w")
        ref_style_var = tk.StringVar(value=str(opt_get("ref_style", "-.")))
        ttk.Combobox(frm, textvariable=ref_style_var, values=["-", "--", "-.", ":"], state="readonly", width=8).grid(row=26, column=1, sticky="w", padx=(6,0))
 
        ttk.Label(frm, text="Width:").grid(row=27, column=0, sticky="w")
        ref_width_var = tk.StringVar(value=str(opt_get("ref_width", 1.25)))
        ttk.Entry(frm, textvariable=ref_width_var, width=8).grid(row=27, column=1, sticky="w", padx=(6,0))
 
        ttk.Label(frm, text="Transparency (alpha):").grid(row=28, column=0, sticky="w")
        ref_alpha_var = tk.StringVar(value=str(opt_get("ref_alpha", 0.9)))
        ttk.Entry(frm, textvariable=ref_alpha_var, width=8).grid(row=28, column=1, sticky="w", padx=(6,0))
 
        ttk.Separator(frm).grid(row=29, column=0, columnspan=2, sticky="ew", pady=(8,8))
 
        # ----- Section: Statistics -----
        ttk.Label(frm, text="Statistics", style=self.options.get("text_style", "TLabel")).grid(row=30, column=0, sticky="w", pady=(0,4))
 
        show_line_stats_var = tk.BooleanVar(value=bool(opt_get("show_line_stats", False)))
        ttk.Checkbutton(frm, text=self.options.get("label_show_line_stats", "Show line statistics"), variable=show_line_stats_var).grid(row=31, column=0, sticky="w")
 
        show_max_in_summary_var = tk.BooleanVar(value=bool(opt_get("show_max_in_summary", False)))
        ttk.Checkbutton(frm, text=self.options.get("label_show_max_in_summary", "Display max in summary"), variable=show_max_in_summary_var).grid(row=31, column=1, sticky="w")
 
        show_custom_message_var = tk.BooleanVar(value=bool(opt_get("show_custom_message", False)))
        ttk.Checkbutton(
            frm,
            text=self.options.get("label_show_custom_message", "Show custom message"),
            variable=show_custom_message_var
        ).grid(row=31, column=3, sticky="w")
 
        # Buttons
        btns = ttk.Frame(frm)
        btns.grid(row=32, column=0, columnspan=2, sticky="e", pady=(12,0))
        def apply_and_refresh():
            # Parse numbers safely
            def _to_int(s, default=None):
                try:
                    return int(str(s).strip())
                except Exception:
                    return default
            def _to_float(s, default=None):
                try:
                    return float(str(s).strip())
                except Exception:
                    return default
 
            # Parse raw values
            raw_x_ticks = x_ticks_var.get()
            raw_y_ticks = y_ticks_var.get()
            x_ticks = _to_int(raw_x_ticks, self.options.get("x_ticks", 5))
            # y_ticks is optional; treat empty as None
            y_ticks = None if str(raw_y_ticks).strip() == "" else _to_int(raw_y_ticks, None)
 
            # Enforce minimum ticks = 3 (with warning)
            if isinstance(x_ticks, int) and x_ticks < 3:
                messagebox.showwarning("Invalid X ticks", "Minimum X ticks is 3. Value has been set to 3.")
                x_ticks = 3
                try:
                    x_ticks_var.set(str(x_ticks))
                except Exception:
                    pass
            if isinstance(y_ticks, int) and y_ticks < 3:
                messagebox.showwarning("Invalid Y ticks", "Minimum Y ticks is 3. Value has been set to 3.")
                y_ticks = 3
                try:
                    y_ticks_var.set(str(y_ticks))
                except Exception:
                    pass
 
            _center_mode_val = str(center_mode_var.get()).lower()
            if _center_mode_val == "x_ref":
                _center_mode_val = "ref"
 
            new_opts = {
                "title": title_var.get(),
                "xlabel": xlabel_var.get(),
                "ylabel": ylabel_var.get(),
                "x_ticks": x_ticks,
                "y_ticks": y_ticks,
                "show_grid": bool(show_grid_var.get()),
                "show_legend": bool(show_legend_var.get()),
                "y_commas": bool(y_commas_var.get()),
                "x_integer_ticks": bool(xint_var.get()),
                "margins": (
                    _to_float(margin_x_var.get(), 0.05) or 0.05,
                    _to_float(margin_y_var.get(), 0.20) or 0.20,
                ),
                "symmetric_x": bool(sym_var.get()),
                "center_mode": _center_mode_val,
                "center_value": _to_float(center_val_var.get(), None),
                # Size
                "custom_size": bool(custom_size_var.get()),
                "width_px": _to_int(width_px_var.get(), None),
                "height_px": _to_int(height_px_var.get(), None),
                "dpi": _to_int(dpi_var.get(), self.options.get("dpi", 100)),
                # Colors
                "main_color": (main_color_var.get().strip() or "#51284F"),
                "accent_color": (accent_color_var.get().strip() or "#0078D4"),
                # Reference line
                "ref_line": bool(ref_line_var.get()),
                "ref_x": _to_float(ref_x_var.get(), None),
                "ref_style": ref_style_var.get(),
                "ref_width": _to_float(ref_width_var.get(), 1.25) or 1.25,
                "ref_alpha": _to_float(ref_alpha_var.get(), 0.9) or 0.9,
                # Statistics
                "show_line_stats": bool(show_line_stats_var.get()),
                "show_max_in_summary": bool(show_max_in_summary_var.get()),
                "show_custom_message": bool(show_custom_message_var.get()),
            }
            self.update_options(new_opts)
            try:
                if hasattr(self, "_title_lbl"):
                    self._title_lbl.configure(style=self.options.get("label_style", "Title.TLabel"))
            except Exception:
                pass
            try:
                self.refresh()
            except Exception as e:
                messagebox.showerror("Customize Chart", f"Failed to refresh chart after applying options:\n{e}")
 
        # --- Auto-apply on edit (debounced) ---
        def _apply_debounced(*_):
            try:
                # cancel any prior scheduled apply
                if hasattr(win, "_apply_after") and win._apply_after is not None:
                    win.after_cancel(win._apply_after)
            except Exception:
                pass
            try:
                win._apply_after = win.after(300, apply_and_refresh)
            except Exception:
                # fallback: apply immediately
                apply_and_refresh()
 
        # Trace changes on all controls to auto-apply
        title_var.trace_add("write", _apply_debounced)
        xlabel_var.trace_add("write", _apply_debounced)
        ylabel_var.trace_add("write", _apply_debounced)
        # x_ticks_var.trace_add("write", _apply_debounced)  # Removed: handled by focus/enter only
        # y_ticks_var.trace_add("write", _apply_debounced)  # Removed: handled by focus/enter only
        show_grid_var.trace_add("write", _apply_debounced)
        show_legend_var.trace_add("write", _apply_debounced)
        y_commas_var.trace_add("write", _apply_debounced)
        ref_line_var.trace_add("write", _apply_debounced)
        ref_x_var.trace_add("write", _apply_debounced)
        margin_x_var.trace_add("write", _apply_debounced)
        margin_y_var.trace_add("write", _apply_debounced)
        sym_var.trace_add("write", _apply_debounced)
        # center mode already has a trace to toggle entry; add another to auto-apply
        center_mode_var.trace_add("write", _apply_debounced)
        center_val_var.trace_add("write", _apply_debounced)
        width_px_var.trace_add("write", _apply_debounced)
        height_px_var.trace_add("write", _apply_debounced)
        dpi_var.trace_add("write", _apply_debounced)
        xint_var.trace_add("write", _apply_debounced)
        main_color_var.trace_add("write", _apply_debounced)
        accent_color_var.trace_add("write", _apply_debounced)
        show_line_stats_var.trace_add("write", _apply_debounced)
        show_max_in_summary_var.trace_add("write", _apply_debounced)
        show_custom_message_var.trace_add("write", _apply_debounced)
 
        # Apply ticks only when user finishes editing (avoid warnings mid-typing)
        try:
            x_ticks_ent.bind("<FocusOut>", lambda e: apply_and_refresh())
            x_ticks_ent.bind("<Return>",   lambda e: apply_and_refresh())
            y_ticks_ent.bind("<FocusOut>", lambda e: apply_and_refresh())
            y_ticks_ent.bind("<Return>",   lambda e: apply_and_refresh())
        except Exception:
            pass
 
        ttk.Button(btns, text="Apply", command=apply_and_refresh).pack(side="right")
        ttk.Button(btns, text="Close", command=win.destroy).pack(side="right", padx=(0,8))
 
        # Lock current natural size and place to the right of the main window
        try:
            win.update_idletasks()
            win.resizable(False, False)
            # widen a bit for the new controls without forcing exact size later
            # let Tk honor requested size from packed widgets
            root = self.winfo_toplevel()
            rx, ry = root.winfo_rootx(), root.winfo_rooty()
            rw = root.winfo_width()
            win.geometry(f"+{rx+rw+10}+{ry+60}")
        except Exception:
            pass
 
        # Make pressing Enter apply changes
        win.bind("<Return>", lambda e: apply_and_refresh())
 
    # ---------- Internals ----------
    def _draw_placeholder(self, msg: str):
        self._ax.clear()
        self._ax.text(
            0.5, 0.5, msg,
            ha="center", va="center",
            fontsize=14, fontweight="bold",
            color="gray", transform=self._ax.transAxes
        )
        self._ax.set_xticks([]); self._ax.set_yticks([])
        self._ax.set_frame_on(False)
        self._canvas.draw()
        try:
            if getattr(self, "_stats_text", None) is not None:
                self._stats_text.remove()
                self._stats_text = None
        except Exception:
            pass
 
 

